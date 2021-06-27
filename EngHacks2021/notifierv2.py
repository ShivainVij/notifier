import urllib, re, os, random
# import urllib.request
from bs4 import BeautifulSoup as soup
# import xlsxwriter
# import pandas as pd
from plyer import notification
from openpyxl import Workbook, load_workbook
# import requests
# import webbrowser
# import cookiejar
import _thread
# from PIL import Image
# import imagehash
from urllib.request import Request, urlopen
import hashlib
# from selenium import webdriver
# import time
from playwright.sync_api import sync_playwright
import pixelmatch
import random





# import pickle
# from selenium import webdriver

# chromeOptions = webdriver.ChromeOptions()

# driver = webdriver.Chrome(options=chromeOptions)
# driver.get("http://mail.google.com")
# input()
# pickle.dump(driver.get_cookies() , open("cookies.pkl","wb"))
# driver.quit

# print("completed getting cookies")
# chromeOptions.headless = True
# driver = webdriver.Chrome(options=chromeOptions)
# driver.get("http://mail.google.com")
# cookies = pickle.load(open("cookies.pkl", "rb"))
# for cookie in cookies:
#     print(cookie)
#     driver.add_cookie(cookie)

# input()
# driver.save_screenshot("/Users/shivainvij/Documents/EngHacks2021/imageVerify.png")

# 




# options = webdriver.ChromeOptions()
# options.add_argument("--no-sandbox")
# options.add_argument('--user-data-dir=/Users/shivainvij/Documents/EngHacks2021')
# options.add_argument("--profile-directory=Profile 1")
# # options.add_argument("--disable-infobars")


# options.headless = True
# driver = webdriver.Chrome(options=options)
# driver.get("https://")
# driver.save_screenshot("/Users/shivainvij/Documents/EngHacks2021/imageVerify.png")


# def ScrapeLoginInitialize(link, driver):
#     driver.get(link)
#     input("Click enter once you have logged in! Please leave the Chrome window open if you would like to continue receiving notifications.")

# def ScrapeHashLoggedIn(link, driver):
#     driver.minimize_window()
#     driver.get(link)
#     driver.minimize_window()
#     time.sleep(10)
#     driver.save_screenshot("imageVerify.png")
#     return str(imagehash.average_hash(Image.open("imageVerify.png")))

import calendar
import time

import math, operator

from PIL import ImageChops, Image
from functools import reduce

def rmsdiff(im1, im2):

    h = ImageChops.difference(im1, im2).histogram()

    # calculate rms
    return math.sqrt(reduce(operator.add,
        map(lambda h, i: h*(i**2), h, range(256))
    ) / (float(im1.size[0]) * im1.size[1]))

def ScrapeHash(link):
    with sync_playwright() as p:
        browser = p.webkit.launch()
        page = browser.new_page()
        page.goto(link)
        time.sleep(15)
        ts = calendar.timegm(time.gmtime())
        randomnum = random.randint(0,9999)
        page.screenshot(path="{}/".format(os.getcwd()) + str(ts) + str(randomnum) + ".png")
        browser.close()

    return "{}/".format(os.getcwd()) + str(ts) + str(randomnum) + ".png"

# ScrapeHash("https://mail.google.com")

    # print("REQUEST")
    # f1 = Request(link, headers={'User-Agent': 'Mozilla/5.0'})
    # f = urlopen(f1)
    # html = f.read()
    # f.close()
    # soup_page = str(soup(html, "html.parser"))
    # soup_split = "\n".join(soup_page.split("\n")[:-7])
    # print(Xx    )
    # return soup_split
    

def DeleteIndikayt():
    os.system('clear')
    wb = load_workbook('{}/notifs.xlsx'.format(os.getcwd()))
    sheet = wb['Sheet1']
    i = 1
    for row in sheet.rows:
        print(str(i) + ". " + str(row[0].value))
        i += 1
    rowEdit = int(input("Indicate the number of the link you would like to stop monitoring: "))
    ws = wb.active
    ws.delete_rows(idx=rowEdit, amount=1)
    wb.save('{}/notifs.xlsx'.format(os.getcwd()))
    

def CreateIndikayt():
    link = input("Link that you would like to monitor for changes: ")
    wb = load_workbook('{}/notifs.xlsx'.format(os.getcwd()))
    
    ws = wb.active
    maxRow = ws.max_row
    ws.cell(row=(maxRow+1), column=1, value=link)
    ws.cell(row=(maxRow+1), column=2, value=ScrapeHash(link))
    wb.save('{}/notifs.xlsx'.format(os.getcwd()))

def notifyUser(site):
    notification.notify(
            #title of the notification,
            title = "New update!",
            #the body of the notification
            message = "Check " + str(site) + " for updates", 
        )

def monitor():
    while True:
        time.sleep(15)
        wb = load_workbook('{}/notifs.xlsx'.format(os.getcwd()))
        sheet = wb['Sheet1']
        ws = wb.active
        i=1
        for row in sheet.rows:
            imgpath = ScrapeHash(str(row[0].value))
            # print(str(row[1].value))
            # print(imgpath)
            if rmsdiff(Image.open(str(row[1].value)), Image.open(imgpath)) > 0.3:
                #Update value
                ws.cell(row=i, column=2, value=imgpath)
                notifyUser(row[0].value)
            # else:
            #     os.remove(imgpath)
            i = i + 1
        wb.save('{}/notifs.xlsx'.format(os.getcwd()))


#Main program starts here

_thread.start_new_thread(monitor, ())

while True:

    print("Welcome to indikayt. Please choose one of the following options: \n")

    print("1. Delete old indikaytor\n")
    print("2. Create new indikaytor\n")

    i = input("Option (1 or 2): ")

    if i not in ['1', '2']:
        print("Oh no! That's not an option! Press Enter to Continue")
        input()
    else:
        if i == '1':
            DeleteIndikayt()
        else:
            CreateIndikayt()
